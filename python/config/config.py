import json
import warnings
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook


class Config:
    def __init__(self):
        self.signal_map     = None       # vbRcSignals sheet
        self.kpi_spec       = None       # KPI sheet
        self.graph_spec     = None       # graphSpec sheet
        self.line_colors    = None       # lineColors sheet
        self.marker_shapes  = None       # markerShapes sheet
        self.calibratables  = {}         # calibratables 

    @classmethod
    def from_json(cls, json_config_path):
        """Create Config object from JSON file."""
        cfg = cls()

        config_struct = cls._load_config(json_config_path)

        # --- SignalMap & KPI & PlotSpec (all in one Excel) ---
        spec_cfg   = config_struct["SignalMap_KPI_PlotSpec"]
        spec_path  = spec_cfg["FilePath"]
        sheet_list = spec_cfg["Sheets"]

        signal_kpi_plot_spec = cls._load_signal_map_kpi_plot_spec(spec_path, sheet_list)

        # Normalize keys for safe access
        sheet_map = {k.lower(): v for k, v in signal_kpi_plot_spec.items()}

        # Assign sheets
        cfg.signal_map      = sheet_map.get("vbrcsignals")
        cfg.graph_spec      = sheet_map.get("graphspec")
        cfg.line_colors     = sheet_map.get("linecolors")
        cfg.marker_shapes   = sheet_map.get("markershapes")
        cfg.kpi_spec        = sheet_map.get("kpi")

        # ✅ normalize signal_map column headers
        if cfg.signal_map is not None:
            cfg.signal_map.columns = cfg.signal_map.columns.str.strip().str.lower()

        # --- Normalize and clean line_colors sheet ---
        if cfg.line_colors is not None and not cfg.line_colors.empty:
            # Normalize column names
            cfg.line_colors.columns = cfg.line_colors.columns.str.strip().str.lower()

            # Filter to only columns that look like RGB numeric data
            rgb_cols = [c for c in cfg.line_colors.columns if c in ["r", "g", "b"]]
            if len(rgb_cols) == 3:
                try:
                    # Extract RGB values, convert to float, clip to [0, 1]
                    cfg.line_colors = (
                        cfg.line_colors[rgb_cols]
                        .astype(float)
                        .clip(0, 1)
                        .to_numpy()
                        .tolist()
                    )
                    print(f"🎨 Loaded {len(cfg.line_colors)} RGB color entries from lineColors.")
                except Exception as e:
                    warnings.warn(f"⚠️ Failed to parse RGB colors from lineColors: {e}")
            else:
                warnings.warn("⚠️ No valid R,G,B columns found in lineColors sheet.")

        # --- Calibration ---
        calib_cfg           = config_struct["Calibration"]
        calib_file          = calib_cfg["FilePath"]
        sheet_defs          = calib_cfg["Sheets"]
        cfg.calibratables   = cls._load_calibratables(calib_file, sheet_defs)

        # Scale PedalPosProIncrease_Th y-values to percentage (multiply by 100)
        if 'PedalPosProIncrease_Th' in cfg.calibratables:
            lim = cfg.calibratables['PedalPosProIncrease_Th']
            if isinstance(lim, dict) and "y" in lim:
                lim["y"] = [v * 100 for v in lim["y"]]

        cfg.graph_spec.columns = cfg.graph_spec.columns.str.strip().str.lower()

        return cfg

    # --------------------
    # Helpers
    # --------------------
    @staticmethod
    def _load_config(file_path):
        file_path = Path(file_path)
        if not file_path.exists():
            raise FileNotFoundError(f"Config file not found: {file_path}")

        with open(file_path, "r", encoding="utf-8") as f:
            params = json.load(f)

        if "SignalMap_KPI_PlotSpec" not in params or "FilePath" not in params["SignalMap_KPI_PlotSpec"]:
            raise ValueError("Missing SignalMap_KPI_PlotSpec.FilePath in config.")
        if "Sheets" not in params["SignalMap_KPI_PlotSpec"]:
            raise ValueError("SignalMap_KPI_PlotSpec.Sheets must be defined in config.")

        if "Calibration" not in params or "FilePath" not in params["Calibration"]:
            raise ValueError("Missing Calibration.FilePath in config.")
        if "Sheets" not in params["Calibration"]:
            raise ValueError("Calibration.Sheets must be defined in config.")

        return params

    @staticmethod
    def _load_signal_map_kpi_plot_spec(file_path, sheet_list):
        """Load vbRcSignals, graphSpec, lineColors, markerShapes, KPI from Excel."""
        file_path = Path(file_path)
        if not file_path.exists():
            raise FileNotFoundError(f"SignalMap_KPI_PlotSpec file not found: {file_path}")

        result = {}
        for sheet_name in sheet_list:
            try:
                # Try to detect the header row by reading first few rows
                preview = pd.read_excel(file_path, sheet_name=sheet_name, nrows=5, header=None)
                header_row = 0
                for i in range(len(preview)):
                    # A valid header row should contain at least 3 non-null cells
                    non_na = preview.iloc[i].notna().sum()
                    if non_na >= 3:
                        header_row = i
                        break

                # Load the actual sheet using the detected header row
                df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row)
                df.columns = df.columns.str.strip().str.lower()
                result[sheet_name] = df

                print(f"✅ Loaded '{sheet_name}' (header at row {header_row+1}) — shape {df.shape}")

            except Exception as e:
                warnings.warn(f'⚠️ Failed to read sheet "{sheet_name}": {e}')
        return result

    @staticmethod
    def _load_calibratables(calib_file, sheet_map):
        file_path = Path(calib_file)
        if not file_path.exists():
            raise FileNotFoundError(f"Calibration file not found: {file_path}")

        wb = load_workbook(file_path, data_only=True)
        calibratables = {}

        for sheet, cal_defs in sheet_map.items():
            if sheet not in wb.sheetnames:
                warnings.warn(f'Sheet "{sheet}" not found in calibration file.')
                continue

            ws = wb[sheet]

            for cal_name, rng in cal_defs.items():
                try:
                    cells = ws[rng]
                    data = [[cell.value for cell in row] for row in cells]
                    df = pd.DataFrame(data).reset_index(drop=True)

                    if df.isna().all().all():
                        calibratables[cal_name] = None
                    elif df.shape[0] == 2:
                        x = [v if v is not None else float("nan") for v in df.iloc[0].tolist()]
                        y = [v if v is not None else float("nan") for v in df.iloc[1].tolist()]
                        calibratables[cal_name] = {"x": x, "y": y}
                    else:
                        calibratables[cal_name] = df

                except Exception as e:
                    warnings.warn(
                        f'Failed to load "{cal_name}" from sheet "{sheet}" range "{rng}": {e}'
                    )
                    calibratables[cal_name] = None

        wb.close()
        return calibratables
